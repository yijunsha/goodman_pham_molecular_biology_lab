# The purpose of this program is to compare the Low coverage regions of the bottom strand to the top strand
# Takes in 2 .xlsx files, one for the top strand, one for the bottom strand, and outputs to a .txt file with summary statistics.
# Very similar to "gap_classification.py"

import openpyxl
import os.path
import sys
import re

# Enter the file name for the strand (either the top or the bottom)
 strandFilename = input('\nEnter the filename for the strand: ')

# Tries to load the file, if it doesn't load, exits with an error message
if os.path.isfile(strandFilename):
	strandWb = openpyxl.load_workbook(strandFilename)
	strSheet = strandWb.worksheets[0]
else:
	sys.exit("Please enter a valid filename.")

# Asks for file name for the comparison file, tries to load, if it doesn't, exits with error
 rfFilename = input('\nEnter the filename for the strand to compare: ')

if os.path.isfile(rfFilename):
	rfWb = openpyxl.load_workbook(rfFilename)
	rfSheet = rfWb.worksheets[0]
else:
	sys.exit("Please enter a valid filename.")

# Creates a summary .txt file to output to at the very end
output = open(strandFilename + "_compare_bottom_to_top.txt", "w") #change

# Constants for each type of region overlap
lowGapCount = 0 # "outside low Coverage region" - the region lies outside of any low coverage regions from the comparison file
lowEndIntoCount = 0 # "low coverage into outside" - partial overlap, with the end of the region falling outside of any low coverage regions
lowGapIntoCount = 0 # "outside into start" - partial overlap, with the start of the region falling outside of any low coverage regions
lowBetweenGapCount = 0 # "between low coverage regions, gap" - overlaps 2 reading frames, with a gap between the low coverage regions
lowBetweenNoGapCount = 0 # "between low coverage regions, no gap" - overlaps 2 reading frames, but the 2 low coverage regions also overlap each other
lowNoNotableCount = 0 # "overlapping low coverage region" - the region is encompassed by a low coverage regions
lowEncompassCount = 0 # "encompassing low coverage region" - the region encompasses a low coverage regions
lowUnsure = 0 # "unsure" - any overlap that falls outside of the above possibilities

# Variables for which column on the excel sheet to find the start and end bp, where to put the output label, and which column holds the "Low Coverage" label
nameCol = 5
lowCover = "Low Coverage"

rfcurrentRow = 2
startCol = 3
endCol = 4
infoCol = 8
identity = ""

# For each row in the excel sheet that contains data, perform the following checks:
for row in range(2, strSheet.max_row + 1):

	if (strSheet.cell(row, nameCol).value != lowCover):
		continue
    # Stores the starting (-50 bp) and ending bp (+ 50 bp) for the region the row in the excel indicates to cover more range
	startBP = strSheet.cell(row, startCol).value - 50
	endBP = strSheet.cell(row, endCol).value + 50

    # If the end bp is less than the start bp of the low coverage region and on the row previous to this row, there are only strings, label "outside low coverage region"
	if (endBP < rfSheet.cell(rfcurrentRow, startCol).value - 50 and
		type(rfSheet.cell(rfcurrentRow - 1, endCol).value) == str):
		strSheet.cell(row, infoCol, 'outside low Coverage region')
		lowGapCount += 1
		continue

    # If the current row that we're looking at is greater than the total number of rows in the comparison file, then count any remainder regions in the strand file as
    # "outside"
	if (rfcurrentRow >= rfSheet.max_row):
		strSheet.cell(row, infoCol, 'outside')
		lowGapCount += 1
		continue

    # If the row previous to the current only has strings, then label the current as "outside low coverage region"
	if (type(rfSheet.cell(rfcurrentRow - 1, endCol).value) == str):
		strSheet.cell(row, infoCol, 'outside low Coverage region')
		lowGapCount += 1
		rfcurrentRow += 1
		continue

    # While the start bp of the region is greater than the end bp of the comparison region, keep moving to the next row in the comparison file until you hit an overlap
	while startBP > rfSheet.cell(rfcurrentRow, endCol).value:
		rfcurrentRow += 1
		if (strSheet.cell(row, nameCol).value != lowCover):
			rfcurrentRow += 1

    # The following if/elif/else finds regions by comparing start and end bps of the original strand file region and the comparison file region and categorizes them accordingly
	if (endBP < rfSheet.cell(rfcurrentRow, startCol).value - 50 and
		startBP > rfSheet.cell(rfcurrentRow - 1, endCol).value + 50 or rfcurrentRow >= rfSheet.max_row):

		identity = "outside low Coverage region"
		lowGapCount += 1

	elif (startBP >= rfSheet.cell(rfcurrentRow, startCol).value - 50 and 
		startBP <= rfSheet.cell(rfcurrentRow, endCol).value + 50 and 
		endBP <= rfSheet.cell(rfcurrentRow + 1, startCol).value - 50 and
		endBP >= rfSheet.cell(rfcurrentRow, endCol).value + 50):

		diff = rfSheet.cell(rfcurrentRow, endCol).value + 50 - startBP
		identity = "low Coverage into outside region by " + str(diff) + "bp"

		lowEndIntoCount += 1

	elif (endBP <= rfSheet.cell(rfcurrentRow, endCol).value + 50 and 
		endBP >= rfSheet.cell(rfcurrentRow, startCol).value - 50 and 
		startBP >= rfSheet.cell(rfcurrentRow - 1, endCol).value + 50 and
		startBP <= rfSheet.cell(rfcurrentRow, startCol).value - 50):

		diff = endBP - rfSheet.cell(rfcurrentRow, startCol).value
		identity = "outside into start by " + str(diff) + "bp"

		lowGapIntoCount += 1

	elif (startBP >= rfSheet.cell(rfcurrentRow, startCol).value - 50 and 
		startBP <= rfSheet.cell(rfcurrentRow, endCol).value + 50 and
		endBP <= rfSheet.cell(rfcurrentRow + 1, endCol).value + 50 and 
		endBP >= rfSheet.cell(rfcurrentRow + 1, startCol).value - 50):

		if rfSheet.cell(rfcurrentRow, endCol).value + 50 > rfSheet.cell(rfcurrentRow + 1, startCol).value - 50:
			identity = "between low coverage regions, no gap"
			lowBetweenNoGapCount += 1
		else:
			identity = "between low coverage regions, gap"
			lowBetweenGapCount += 1

	elif (startBP <= rfSheet.cell(rfcurrentRow, endCol).value + 50 and 
		startBP >= rfSheet.cell(rfcurrentRow, startCol).value - 50 and
		endBP <= rfSheet.cell(rfcurrentRow, endCol).value + 50 and 
		endBP >= rfSheet.cell(rfcurrentRow, startCol).value - 50):

		identity = "overlapping low coverage region"

		lowNoNotableCount += 1

	elif ((startBP <= rfSheet.cell(rfcurrentRow, startCol).value - 50 and 
		endBP >= rfSheet.cell(rfcurrentRow, endCol).value + 50) or
		(startBP <= rfSheet.cell(rfcurrentRow + 1, startCol).value - 50 and 
		endBP >= rfSheet.cell(rfcurrentRow + 1, endCol).value + 50)):

		identity = "encompassing low coverage region"

		lowEncompassCount += 1

	else:
		identity = "unsure"

		lowUnsure += 1

    # write the identity into the excel sheet's info columns
	strSheet.cell(row, infoCol, identity)

# In the output file, write the summary statistics for the low coverage regions for each type of region overlap
output.write("Summary statistics:\n")
output.write("For Low Coverage Regions:\n")
output.write("	outside low Coverage region: " + str(lowGapCount)  + "\n")
output.write("	low Coverage into outside region: " + str(lowEndIntoCount) + "\n")
output.write("	outside into start: " + str(lowGapIntoCount) + "\n")
output.write("	between low coverage regions, gap: " + str(lowBetweenGapCount) + "\n")
output.write("	between low coverage regions, no gap: " + str(lowBetweenNoGapCount) + "\n")
output.write("	overlapping low coverage region: " + str(lowNoNotableCount) + "\n")
output.write("	encompassing low coverage region: " + str(lowEncompassCount) + "\n")
output.write("	Unsure: " + str(lowUnsure) + "\n")
output.write("	Total Low Coverage: " + str(lowGapCount + lowEndIntoCount + lowGapIntoCount + 
	lowBetweenGapCount + lowBetweenNoGapCount + lowNoNotableCount + lowEncompassCount + lowUnsure) + "\n")

# Save the edited excel file as a new .xlsx file
strandWb.save(strandFilename + "_compare_bottom_to_top.xlsx") #change
