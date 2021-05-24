# The purpose of this program is to compare the low/high coverage regions in each strand to existing known annotations (eg. CDS, ncRNA, biological regions)
# Both input files need to be in .xlsx format. The program will revise the excel sheet containing the strand regions.
# A summary output file will also be generated as a .txt file that contains the raw values for each type of region.

import openpyxl
import os.path
import sys
import re

# Enter the file name for the strand (either the top or the bottom)
strandFilename = input('\nEnter the filename for the strand: ')

# Tries to load the file, if it doesn't load, exits with an error message
if os.path.isfile(strandFilename):
	strandWb = openpyxl.load_workbook(strandFilename)
	strSheet = strandWb.worksheets[0]
else:
	sys.exit("Please enter a valid filename.")

# Asks for file name for the comparison regions (annotations), tries to load, if it doesn't, exits with error
rfFilename = input('\nEnter the filename for the reading frames to compare: ')

if os.path.isfile(rfFilename):
	rfWb = openpyxl.load_workbook(rfFilename)
	rfSheet = rfWb.worksheets[0]
else:
	sys.exit("Please enter a valid filename.")

# Creates a summary .txt file to output to at the very end
output = open(strandFilename + "_summary.txt", "w") #change

# Constants for each type of region overlap
lowGapCount = 0 # "gap in reading frames" - the region lies outside of any annotation regions
lowEndIntoCount = 0 # "end into gap" - partial overlap, with the end of the region falling outside of any annotation regions
lowGapIntoCount = 0 # "gap into start" - partial overlap, with the start of the region falling outside of any annotation regions
lowBetweenGapCount = 0 # "between reading frames, gap" - overlaps 2 reading frames, with a gap between the annotation regions
lowBetweenNoGapCount = 0 # "between reading frames, no gap" - overlaps 2 reading frames, but the 2 annotation regions also overlap each other
lowNoNotableCount = 0 # "no notable" - the region is encompassed by an annotation region
lowEncompassCount = 0 # "encompassing reading frame" - the region encompasses an annotation region
lowUnsure = 0 # "unsure" - any overlap that falls outside of the above possibilities

highGapCount = 0
highEndIntoCount = 0
highGapIntoCount = 0
highBetweenGapCount = 0
highBetweenNoGapCount = 0
highNoNotableCount = 0
highEncompassCount = 0
highUnsure = 0

# Variables for which column on the excel sheet to find the start and end bp, where to put the output label, and which column holds the "Low Coverage"/"High Coverage" labels
nameCol = 5
lowCover = "Low Coverage"
highCover = "High Coverage"

rfcurrentRow = 2
startCol = 3
endCol = 4
infoCol = 8
identity = ""

# For each row in the excel sheet that contains data, perform the following checks:
for row in range(2, strSheet.max_row + 1):

    # Stores the starting and ending bp for the region the row in the excel indicates
	startBP = strSheet.cell(row, startCol).value
	endBP = strSheet.cell(row, endCol).value

    # If the end bp is less than the start bp of the annotation region and on the row previous to this row, there are only strings, label this a gap in reading frames
	if (endBP < rfSheet.cell(rfcurrentRow, startCol).value and
		type(rfSheet.cell(rfcurrentRow - 1, endCol).value) == str):
		strSheet.cell(row, infoCol, 'gap in reading frames')
        
        # Checks to see if it was a low coverage or high coverage region and adds it to the running total
        if (strSheet.cell(row, nameCol).value == lowCover):
			lowGapCount += 1
		else:
			highGapCount += 1
		continue

    # If the current row that we're looking at is greater than the total number of rows in the annotation file, then count any remainder regions in the strand file as
    # "gap in reading frames"
	if (rfcurrentRow >= rfSheet.max_row):
		strSheet.cell(row, infoCol, 'gap in reading frames')
        
		if (strSheet.cell(row, nameCol).value == lowCover):
			lowGapCount += 1
		else:
			highGapCount += 1
		continue

    # While the start bp of the region is greater than the end bp of the annotation region, keep moving to the next row in the annotation file until you hit an overlap
	while startBP > rfSheet.cell(rfcurrentRow, endCol).value:
		rfcurrentRow += 1
		if (rfcurrentRow >= rfSheet.max_row):
			strSheet.cell(row, infoCol, 'gap in reading frames')
		if (strSheet.cell(row, nameCol).value == lowCover):
			lowGapCount += 1
		else:
			highGapCount += 1
		break

    # If the end bp is less than the start bp of the annotation region and the start bp is greater than the previous annotation region's end bp, then label as
    # "gap in reading frames"
	if (endBP < rfSheet.cell(rfcurrentRow, startCol).value and
		startBP > rfSheet.cell(rfcurrentRow - 1, endCol).value or rfcurrentRow >= rfSheet.max_row):

		identity = "gap in reading frames"

		if (strSheet.cell(row, nameCol).value == lowCover):
			lowGapCount += 1
		else:
			highGapCount += 1

    # If the start bp is greater than the start bp of the annotation and less than the end bp of the annotation,
    # and also the end bp is less than the next annotation region's start bp and greater than the current annotation's end bp, then label it as
    # "end into gap"
	elif (startBP >= rfSheet.cell(rfcurrentRow, startCol).value and 
		startBP <= rfSheet.cell(rfcurrentRow, endCol).value and 
		endBP <= rfSheet.cell(rfcurrentRow + 1, startCol).value and
		endBP >= rfSheet.cell(rfcurrentRow, endCol).value):

        # figure out how many bp the overlap is
		diff = rfSheet.cell(rfcurrentRow, endCol).value - startBP
		identity = "end into gap by " + str(diff) + "bp"

		if (strSheet.cell(row, nameCol).value == lowCover):
			lowEndIntoCount += 1
		else:
			highEndIntoCount += 1

    # If the end bp is less than the end bp of the annotation and greater than the start bp of the annotation,
    # and also the start bp is greater than the previous annotation region's end bp and less than the current annotation's start bp, then label it as
    # "gap into start"
	elif (endBP <= rfSheet.cell(rfcurrentRow, endCol).value and 
		endBP >= rfSheet.cell(rfcurrentRow, startCol).value and 
		startBP >= rfSheet.cell(rfcurrentRow - 1, endCol).value and
		startBP <= rfSheet.cell(rfcurrentRow, startCol).value):

        # figure out how many bp the overlap is
		diff = endBP - rfSheet.cell(rfcurrentRow, startCol).value
		identity = "gap into start by " + str(diff) + "bp"

		if (strSheet.cell(row, nameCol).value == lowCover):
			lowGapIntoCount += 1
		else:
			highGapIntoCount += 1

    # If the start bp is greater than the start bp of the annotation and less than the end bp of the annotation,
    # and also the end bp is less than the next annotation region's end bp and greater than the next annotation's start bp, then label it as
    # "between reading frames"
	elif (startBP >= rfSheet.cell(rfcurrentRow, startCol).value and 
		startBP <= rfSheet.cell(rfcurrentRow, endCol).value and
		endBP <= rfSheet.cell(rfcurrentRow + 1, endCol).value and 
		endBP >= rfSheet.cell(rfcurrentRow + 1, startCol).value):

        # check to see if there's a gap b/w the annotation regions themselves and label accordingly
		if rfSheet.cell(rfcurrentRow, endCol).value > rfSheet.cell(rfcurrentRow + 1, startCol).value:
			identity = "between reading frames, no gap"

			if (strSheet.cell(row, nameCol).value == lowCover):
				lowBetweenNoGapCount += 1
			else:
				highBetweenNoGapCount += 1
		else:
			identity = "between reading frames, gap"
			if (strSheet.cell(row, nameCol).value == lowCover):
				lowBetweenGapCount += 1
			else:
				highBetweenGapCount += 1

    # If the start bp is less than the end bp of the annotation and greater than the start bp of the annotation,
    # and also the end bp is less than the annotation region's end bp and greater than the annotation's start bp, then label it as
    # "no notable"
	elif (startBP <= rfSheet.cell(rfcurrentRow, endCol).value and 
		startBP >= rfSheet.cell(rfcurrentRow, startCol).value and
		endBP <= rfSheet.cell(rfcurrentRow, endCol).value and 
		endBP >= rfSheet.cell(rfcurrentRow, startCol).value):

		identity = "no notable"

		if (strSheet.cell(row, nameCol).value == lowCover):
			lowNoNotableCount += 1
		else:
			highNoNotableCount += 1

    # If the start bp is less than the start bp of the annotation and the end bp is greater than the end bp of the annotation,
    # or the start bp is less than the next annotation region's start bp and the end bp is greater than the next annotation's end bp, then label it as
    # "encompassing reading frame"
	elif ((startBP <= rfSheet.cell(rfcurrentRow, startCol).value and 
		endBP >= rfSheet.cell(rfcurrentRow, endCol).value) or
		(startBP <= rfSheet.cell(rfcurrentRow + 1, startCol).value and 
		endBP >= rfSheet.cell(rfcurrentRow + 1, endCol).value)):

		identity = "encompassing reading frame"

		if (strSheet.cell(row, nameCol).value == lowCover):
			lowEncompassCount += 1
		else:
			highEncompassCount += 1
            
    # anything else, label as "unsure"
	else:
		identity = "unsure"

		if (strSheet.cell(row, nameCol).value == lowCover):
			lowUnsure += 1
		else:
			highUnsure += 1

    # write the identity into the excel sheet's info column
	strSheet.cell(row, infoCol, identity)

# In the output file, write the summary statistics for the low/high coverage regions for each type of region overlap
output.write("Summary statistics:\n")
output.write("For Low Coverage Regions:\n")
output.write("	Gap in Reading frame: " + str(lowGapCount)  + "\n")
output.write("	End into Gap: " + str(lowEndIntoCount) + "\n")
output.write("	Gap into Start: " + str(lowGapIntoCount) + "\n")
output.write("	Between reading frames w Gap: " + str(lowBetweenGapCount) + "\n")
output.write("	Between reading frames w/o Gap: " + str(lowBetweenNoGapCount) + "\n")
output.write("	No Notable: " + str(lowNoNotableCount) + "\n")
output.write("	Encompassing reading frame: " + str(lowEncompassCount) + "\n")
output.write("	Unsure: " + str(lowUnsure) + "\n")
output.write("	Total Low Coverage: " + str(lowGapCount + lowEndIntoCount + lowGapIntoCount + 
	lowBetweenGapCount + lowBetweenNoGapCount + lowNoNotableCount + lowEncompassCount + lowUnsure) + "\n")

output.write("Summary statistics:\n")
output.write("For High Coverage Regions:\n")
output.write("	Gap in Reading frame: " + str(highGapCount)  + "\n")
output.write("	End into Gap: " + str(highEndIntoCount) + "\n")
output.write("	Gap into Start: " + str(highGapIntoCount) + "\n")
output.write("	Between reading frames w Gap: " + str(highBetweenGapCount) + "\n")
output.write("	Between reading frames w/o Gap: " + str(highBetweenNoGapCount) + "\n")
output.write("	No Notable: " + str(highNoNotableCount) + "\n")
output.write("	Encompassing reading frame: " + str(highEncompassCount) + "\n")
output.write("	Unsure: " + str(highUnsure) + "\n")
output.write("	Total High Coverage: " + str(highGapCount + highEndIntoCount + highGapIntoCount + 
	highBetweenGapCount + highBetweenNoGapCount + highNoNotableCount + highEncompassCount + highUnsure) + "\n")

# Save the edited excel file as a new .xlsx file
strandWb.save(strandFilename + "_classified.xlsx") #change
