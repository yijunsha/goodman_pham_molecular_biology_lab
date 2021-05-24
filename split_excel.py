# This program was made to split the start and end bp into different columns from excel sheets extracted from the CLC Workbench program
# Data was originally presented as "start..end" and were split into separate start columns and end columns

import openpyxl
import os.path
import sys
import re

# Read in file and check if it's valid, if not, then terminate the program
filename = input('\nEnter the filename: ')
if os.path.isfile(filename):
	wb = openpyxl.load_workbook(filename)
	sheet = wb.worksheets[0]
else:
	sys.exit("Please enter a valid filename.")

# insert 2 new columns to hold the start bp and end bp
sheet.insert_cols(2)
sheet.insert_cols(4)

sheet.cell(1, 2, 'read direction')
sheet.cell(1, 3, 'start bp')
sheet.cell(1, 4, 'end bp')


currentCol = 3

# for each row in the file, check to see if there are "c" or "j"'s in the info file, if so, avoid them when splitting and take out any parentheses
for row in range(1, sheet.max_row + 1):
	cellValue = sheet.cell(row, currentCol).value
	if cellValue[0] == 'c' or cellValue[0] == 'j':
		split = cellValue.split('(')
		sheet.cell(row, currentCol - 1, split[0])
		split[1] = split[1].replace(')','')
		sheet.cell(row, currentCol, split[1])
	cellValue = sheet.cell(row, currentCol).value
 
    # split the value around ".." and then put the value to the right as the end bp and the value to the left as start bp
	ellipseSplit = cellValue.split('..')
	if (len(ellipseSplit) > 1):
		startBP = int(ellipseSplit[0])
		endBP = int(ellipseSplit[1])
		sheet.cell(row, currentCol, startBP)
		sheet.cell(row, currentCol + 1, endBP)

# save the xlsx as a new file and append "_split" to the name
wb.save(filename + "_split.xlsx")
